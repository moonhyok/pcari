from django.contrib import admin

from django.http import HttpResponse
from django.core import serializers
from django.contrib.contenttypes.models import ContentType
from django.http import HttpResponseRedirect

# Register your models here.
from pcari.models import QuantitativeQuestion, QualitativeQuestion, Rating, Comment, UserProgression, GeneralSetting, FlaggedComment, UserData


# admin.site.register(QuantitativeQuestion)
admin.site.register(QualitativeQuestion)
# admin.site.register(Rating)
# admin.site.register(Comment)
# admin.site.register(UserProgression)

def export_comment_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=mymodel.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"User"),
        smart_str(u"Average_score"),
        smart_str(u"Number_rated"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.user),
            smart_str(obj.average_score),
            smart_str(obj.number_rated),
        ])
    return response

export_comment_csv.short_description = u"Export as CSV"

def export_comment_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=malasakit_data.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Malasakit Data"

    row_num = 0

    columns = [
        (u"User", 15),
        (u"Average_score", 70),
        (u"Number_rated", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.user,
            obj.average_score,
            obj.number_rated,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

def export_question_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=mymodel.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"User"),
        smart_str(u"Average_score"),
        smart_str(u"Number_rated"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.user),
            smart_str(obj.average_score),
            smart_str(obj.number_rated),
        ])
    return response

export_question_csv.short_description = u"Export as CSV"

def export_question_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=malasakit_data.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Malasakit Data"

    row_num = 0

    columns = [
        (u"User", 15),
        (u"Average_score", 70),
        (u"Number_rated", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.user,
            obj.average_score,
            obj.number_rated,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response


export_question_xlsx.short_description = u"Export as XLSX"


class GeneralSettingAdmin(admin.ModelAdmin):
    actions = None

admin.site.register(GeneralSetting, GeneralSettingAdmin)


def flag_comment(modeladmin, request, queryset):
	for obj in queryset:
		f = FlaggedComment(user=obj.user,comment=obj.comment,tagalog_comment=obj.tagalog_comment,date=obj.date,average_score=obj.average_score,number_rated=obj.number_rated,tag=obj.tag)
		f.save()
		obj.delete()

flag_comment.short_description = u"Flag Comment"

def unflag_comment(modeladmin, request, queryset):
	for obj in queryset:
		c = Comment(user=obj.user,comment=obj.comment,tagalog_comment=obj.tagalog_comment,date=obj.date,average_score=obj.average_score,number_rated=obj.number_rated,tag=obj.tag)
		c.save()
		obj.delete()

unflag_comment.short_description = u"Unflag Comment"


class CommentAdmin(admin.ModelAdmin):
    list_display = ['user', 'comment', 'tagalog_comment', 'average_score', 'number_rated', 'tag']
    list_editable = ['comment','tagalog_comment']
    ordering = ['user']
    actions = [flag_comment,export_comment_csv, export_comment_xlsx]

admin.site.register(Comment, CommentAdmin)

class FlaggedCommentAdmin(admin.ModelAdmin):
    list_display = ['user', 'comment', 'tagalog_comment', 'average_score', 'number_rated', 'tag']
    ordering = ['user']
    actions = [unflag_comment,export_comment_csv, export_comment_xlsx]

admin.site.register(FlaggedComment, FlaggedCommentAdmin)


class QuantitativeQuestionAdmin(admin.ModelAdmin):
    list_display = ['question', 'tagalog_question', 'average_score']
    ordering = ['qid']
    actions = [export_question_csv, export_question_xlsx]

admin.site.register(QuantitativeQuestion, QuantitativeQuestionAdmin)

# class UserProgressionAdmin(admin.ModelAdmin):
# 	a = 10
# 	list_display = [a]

admin.site.register(UserData)
# admin.site.register(UserProgression, UserProgressionAdmin)
